import requests 
import openpyxl
BOT_TOKEN = '6713141695:AAGMXTZ5ZYMWeLG4B67euvpp0ylxZbxxy3o'
workbook = openpyxl.load_workbook('Ll.xlsx')
sheet = workbook.active
max_row = sheet.max_row
for row in range(1, 2, max_row + 1):
    name = sheet.cell(row=row, column=1).value
    student_id = sheet.cell(row=row, column=2).value
    if student_id:
        telegram_url = f'https://api.telegram.org/bot{BOT_TOKEN}/sendMessage'
        message = f'Hello, {name}!'
        params = {
            'chat_id': student_id,
            'text': message
        }
        response = requests.get(telegram_url, params=params)
        print(f'Message sent to student {name} with ID {student_id}')
    else:
        print(f'Student {name} does not have an ID')
workbook.save('Ll.xlsx')